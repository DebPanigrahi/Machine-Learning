import os, struct
import pandas as pd
from pylab import *
from numpy import *
import scipy.sparse as sparse
import scipy.linalg as LA
import scipy.stats
import openpyxl as pyxl
from random import choices
from multiprocessing.dummy import Pool as ThreadPool

# conda install -c anaconda graphviz
# conda install -c anaconda python-graphviz
# import graphviz

def arrayToDataFrame(a, **kwargs):
    if type in kwargs:
        df = pd.DataFrame(a, columns=kwargs[type])
    else:
        df = pd.DataFrame(a)
    return df

def freq_count(a, **kwargs):
    freq = scipy.stats.itemfreq(a)
    if 'classes' in kwargs:
        cls = kwargs['classes']
        x = pd.Series(zeros(cls.shape, dtype=int), index=cls)
    else:
        x = pd.Series(zeros(freq.shape[0]), index=freq[:, 0], dtype=int)

    for l in freq: x[l[0]] = l[1]
    return x

def series_to_tuple(S):
    return str([(int(idx), val) for idx, val in S.iteritems()])

def series_to_dataframe(a):
    # display all columns
    pd.set_option('display.max_columns', None)
    for i, srs in enumerate(a):
        if i==0:
            z = pd.DataFrame(srs, dtype=srs.dtype)
        else:
            z.insert(i, column=srs.name, value=srs)
            #z[[srs.name]] = z[[srs.name]].astype(srs.dtype)
    return z.transpose()


## Read/Write to xls files

class Excel:
    def __init__(self, file, **kwargs):
        self.path = file

        if os.path.isfile(file):
            self.xl = pyxl.load_workbook(file)
        else:
            self.xl = pyxl.Workbook()

        if 'sheet' in kwargs:
            self.open_sheet(kwargs['sheet'])
        return


    def open_sheet(self, sheet_name, **kwargs):
        self.sheet_name = sheet_name
        if sheet_name not in self.xl.sheetnames:
            self.sheet = self.xl.create_sheet(title=sheet_name)
        else:
            self.sheet = self.xl.get_sheet_by_name(sheet_name)
        return

    def read(self, **kwargs):
        if 'sheet' in kwargs:
            self.open_sheet(kwargs['sheet'])
        data = pd.read_excel(self.path, sheetname=self.sheet_name, header=None)
        return data

    def write_array(self, value, dest, **kwargs):
        if type(value) in (list, tuple, pd.DataFrame):
            value = array(value)

        if 'sheet' in kwargs:
            self.open_sheet(kwargs['sheet'])


        if type(value) in (float, int):
            self.sheet.cell(row=dest[0], column=dest[1]).value = value
        elif len(value.shape)==1:
            for i, val in enumerate(value):
                self.sheet.cell(row=dest[0], column=dest[1]+i).value = val
        elif len(value.shape)==2:
            for i, valr in enumerate(value):
                for j, valc in enumerate(valr):
                    self.sheet.cell(row=dest[0]+i, column=dest[1]+j).value = str(valc)

        if ('save' in kwargs) and kwargs['save']==True:
            self.xl.save(self.path)

    def save(self):
        self.xl.save(self.path)


class ClassifiedData:
    # X is the training set and Y is the class label
    def __init__(self, Xs, Ys, **kwargs):
        # transform info DataFrame if its not already one
        if type(Xs) == np.ndarray:
            # Here labels can only be N X 1 dimension
            self.Xs = arrayToDataFrame(Xs, 'features', kwargs)
            self.Ys = arrayToDataFrame(Ys, 'labels_categories', kwargs)
        else:
            self.Xs = Xs
            self.Ys = Ys

        self.features = list(self.Xs)
        self.label_categories = list(self.Ys)
        self.N, self.d = self.Xs.shape
        self.class_labels = {}
        for lc in self.label_categories:
            self.class_labels[lc] = unique(self.Ys[lc])

        if 'pca_dimension' in kwargs:
            self.pca(pca_dimension=kwargs['pca_dimension'])
        elif 'pca_accuracy' in kwargs:
            self.pca(pca_accuracy=kwargs['pca_accuracy'])
        elif 'kesler_encoding' in kwargs and kwargs['kesler_encoding']==True:
            self.kesler_encoding()
        return

    def pca(self, **kwargs):
        μ = np.mean(self.Xs)
        Z = self.Xs - μ
        N, d = self.Xs.shape
        C = np.cov(Z, ddof=1, rowvar=False)
        λ, V = LA.eigh(C)
        # λ is 1Xd vector with eighen values. corresponding eighen vector for λ[i] is V[:,i]
        # print(np.dot(transpose(V), V))
        λ = λ / sum(λ)
        acc = np.array([sum(λ[i:]) for i in range(d)])
        Vr = V
        if 'pca_dimension' in kwargs:
            Vr = V[:, -int(kwargs['pca_dimension']):]
            print("Accuracy", acc[-kwargs['pca_dimension'] - 1])
        elif 'pca_accuracy' in kwargs:
            for i in range(d):
                if (acc[-i - 1] > kwargs['pca_accuracy']):
                    Vr = V[:, -i - 1:]
                    print("Dimension reduction ", (1- (i + 1)/d)*100)
                    break
        self.Xs = pd.DataFrame(dot(Z, Vr), index=self.Xs.index)
        self.features = list(self.Xs)
        self.Vr, self.μ, self.λ = Vr, μ, λ
        return

    def pca_validate(self, X):
        I = dot(self.Vr, transpose(self.Vr))
        Z = X - self.μ
        P = dot(Z, self.Vr)
        RZ = dot(P, transpose(self.Vr))
        RX = RZ + self.μ
        return Z, P, RZ, RX

    def kesler_encoding(self):
        # used for linear classifier
        self.kesler_map = {}
        self.Zs = {}
        for lc in self.label_categories:
            self.kesler_map[lc] = {}
            class_labels = unique(self.Ys[lc])
            class_to_int = {}
            for i, cl in enumerate(class_labels):
                self.kesler_map[lc][i] = cl
                class_to_int[cl] = i
            bits_array = (-1 * ones((self.N, len(class_labels)), dtype=int))
            bits_array[arange(self.N), array(self.Ys[lc].map(class_to_int))] = 1
            self.Zs[lc] = pd.DataFrame(bits_array)
        return

def pdf(X, μ, Cov):
    Z = X-μ
    d = Cov.shape[0]
    exponent = -0.5*np.matmul(np.matmul(Z, np.linalg.inv(Cov)), Z)
    factor = 1.0/(((2.0*math.pi)**(d/2))*math.sqrt(np.linalg.det(Cov)))
    pdfx = factor*np.exp(exponent)
    return pdfx


class Histogram:
    def __init__(self, data, **kwargs):
        if 'number_of_bins' in kwargs:
            B = kwargs['number_of_bins']
        else:
            B = round(math.log(data.Xs.shape[0], 2))

        if type(B) == (int):
            B = B * ones(data.d, dtype=int32)

        self.data = data
        self.B = B
        self.Dmax = Dmax = data.Xs.max()
        self.Dmin = Dmin = data.Xs.min()
        width = Dmax - Dmin
        for i, v in enumerate(width):
            if v == 0:
                width[i] = 1
                B[i] = 1

        self.bins = {}
        for class_label in data.class_labels:
            self.bins[class_label] = zeros(B, dtype=int32)
            bin_indices = ((data.Xs.loc[[class_label]] - Dmin) * B / width).astype('int32')
            for i, feature in enumerate(data.features):
                bin_indices[i].replace(B[i], B[i] - 1, inplace=True)
            # could not get it to work for multidimension: revisit later. May be a histogram-dataframe would work
            for i, idx in bin_indices.iterrows():
                self.bins[class_label][idx[0], idx[1]] += 1

    def classify(self, X):
        N_bin = {}
        N_bin_total = 0
        for class_label in self.data.class_labels:
            bin_index = ((X - self.Dmin) * self.B / (self.Dmax-self.Dmin)).astype('int32')
            for i, feature in enumerate(self.data.features):
                if (bin_index[i] == self.B[i]): bin_index[i] -= 1
            N_bin[class_label] = self.bins[class_label][bin_index[0], bin_index[1]]
            N_bin_total += N_bin[class_label]

        probability = {}
        for class_label in self.data.class_labels:
            probability[class_label] = N_bin[class_label] / N_bin_total

        classified_label, classification_probability = max(probability.items(), key=lambda x: x[1])
        return classified_label, classification_probability


class NormalBayesian:
    def __init__(self, data, **kwargs):
        self.data = data
        covariance = {}
        μ = {}
        N = {}
        for lc in data.label_categories:
            covariance[lc] = {}
            μ[lc] = {}
            N[lc] = {}
            for class_label in data.class_labels[lc]:
                Xs = data.Xs[data.Ys[lc]==class_label]
                covariance[lc][class_label] = np.cov(Xs, ddof=1, rowvar=False)
                μ[lc][class_label] = mean(Xs)
                N[lc][class_label] = len(Xs)

        self.covariance = covariance
        self.N = N
        self.μ = μ
        return

    def classify(self, X):
        classified_label = {}
        classification_probability = {}
        for lc in data.label_categories:
            class_pdf = {}
            probability = {}
            for class_label in self.data.class_labels[lc]:
                class_pdf[class_label] = pdf(X, self.μ[lc][class_label], self.covariance[lc][class_label])

            total_probablity = sum(self.N[lc][class_label]*class_pdf[class_label] for class_label in self.data.class_labels[lc])
            for class_label in self.data.class_labels[lc]:
                probability[class_label] = self.N[lc][class_label]*class_pdf[class_label]/total_probablity

            classified_label[lc], classification_probability[lc] = max(probability.items(), key=lambda x: x[1])
        return classified_label, classification_probability


class LinearMinSquare:
    def __init__(self, data):
        x0 = ones(data.N)
        self.data = data
        self.W = {}
        for lc in data.label_categories:
            self.W[lc] = matmul(pinv(column_stack((x0, data.Xs))), data.Zs[lc])
        return

    def classify(self, X):
        # inject dummy probability of 1
        label = {}
        probability = {}
        for lc in data.label_categories:
            T = matmul(insert(array(X), 0, 1), self.W[lc])
            label[lc] = self.data.kesler_map[lc][argmax(T)]
            probability[lc] = 1
        return label, probability


class partitionGini:
    def __init__(self, Y):
        class_labels = array([-1, 1])
        self.Y = Y
        self.NL = freq_count([], classes=class_labels)
        self.N = freq_count(self.Y, classes=class_labels)
        self.NR = self.N.copy()
        self.set_impurity()
        self.pNode = self.N/self.N.sum()
        if self.pNode[-1] > 0.5:
            self.label = -1
        else:
            self.label = 1
        self.probability = self.pNode[self.label]
        self.index = -1
        return

    def update(self, y, **kwargs):
        self.NL[y] += 1
        self.NR[y] -= 1
        self.set_impurity()
        self.index += 1

        return

    def __str__(self):
        return str("LEFT: "+series_to_tuple(self.NL)+", RIGHT: "+series_to_tuple(self.NR)+", impurity = "+str(self.impurity))

    def set_impurity(self):
        self.impurity = 0
        if (self.NR != 0).all():
            self.impurity = (self.NR[-1]*self.NR[1])/(self.NR[-1]+self.NR[1])
        if (self.NL != 0).all():
            self.impurity += (self.NL[-1]*self.NL[1])/(self.NL[-1]+self.NL[1])

        self.impurity /= self.N.sum()
        return


class Node:
    def __init__(self, type, index, count, **kwargs):
        # stores both leaf and node
        self.type = type
        self.index = index
        self.count = count
        self.depth = kwargs['depth']
        self.label = kwargs['label']
        self.probability = kwargs['probability']
        if type == 'rule':
            self.feature = kwargs['feature']
            self.threshold = kwargs['threshold']
            self.left = int(index * 2)
            self.right = int(index * 2 + 1)
        return

    def __str__(self):
        cnt = str([(val, int(idx)) for idx, val in self.count.iteritems()])
        if self.type == 'rule':
            return str(
                "__" * self.depth + str(self.index) + " rule , feature=" + str(self.feature) + ", threshold=" + str(
                    self.threshold) + ", count=" + cnt)
        else:
            return str(
                "__" * self.depth + str(self.index) + " LEAF , label=" + str(self.label) + ", probability=" + str(
                    self.probability) + ", count=" + cnt)


class decisionTree:
    def stump(self, X, Y, **kwargs):
        # for a given feature it identifies optimal splitting
        S = X.sort_values()
        if self.method == 'gini':
            p = partitionGini(Y)
        else:
            p = partitionGini(Y)

        stump_cut = pd.Series(ones(len(X), dtype=int), index=X.index, name='cut')
        impurities = pd.Series(index=X.index, name='impurity')
        Iopt = p.impurity
        I0 = p.impurity
        t = -1.001 * abs(S[S.index[0]])

        icut = -1
        for i, idx in enumerate(S.index[:-1]):
            p.update(Y[idx])
            if p.impurity < Iopt:
                Iopt = p.impurity
                t = (S[idx] + S[S.index[i + 1]]) / 2.0
                # t = S[idx]
                icut = i

            #print(p)
            if 'print' in kwargs and kwargs['print']:
                print(p)
                impurities[idx] = p.impurity


        for i, idx in enumerate(S.index[:-1]):
            if i < icut:
                stump_cut[idx] = -1
            else:
                break

        if 'print' in kwargs and kwargs['print']:
            print(series_to_dataframe([S,Y, impurities, stump_cut]))
            print("threshold=", t, "\n edge=", I0-Iopt)
            print(X.name)
        return t, stump_cut, I0 - Iopt

    def cut(self, X, Y, index, depth):
        features = self.data.features
        best_edges = zeros(len(features))
        best_cuts = zeros(len(features))
        stump_cuts = pd.DataFrame(ones(X.shape, dtype=int), columns=features, index=X.index)
        stump_pool = ThreadPool(8)
        stump_input = []
        for i, feature in enumerate(features):
            stump_input.append((X[feature], Y))
        #stump_results = stump_pool.starmap(self.stump, stump_input)
        #stump_pool.close()
        #stump_pool.join()
        #for i, res in enumerate(stump_results):
        #    best_cuts[i], stump_cuts[features[i]], best_edges[i] = res

        for i, feature in enumerate(features):
            if i == 3 and depth == 2: print_stump = True
            else: print_stump = False
            best_cuts[i], stump_cuts[features[i]], best_edges[i] = self.stump(X[feature], Y, print=print_stump)

        idx = argmax(best_edges)
        # if index > 3 and index < 6:
        #     print(idx, features[idx], best_edges)
        counts = freq_count(stump_cuts[features[idx]])
        # just to get counts
        if self.method == 'gini':
            p = partitionGini(Y)
        else:
            p = partitionGini(Y)
        if -1 in counts:
            rule = Node('rule', index, p.N, depth=depth + 1, label=p.label, probability=p.probability, feature=features[idx], threshold=best_cuts[idx])
            self.tree[index] = rule
            print(rule)
            # print(X[stump_cuts[features[idx]]==-1][features[idx]], Y[stump_cuts[features[idx]]==-1])

            pool = ThreadPool(2)
            args = [(X[stump_cuts[features[idx]] == -1], Y[stump_cuts[features[idx]] == -1], rule.left, depth + 1),
                    (X[stump_cuts[features[idx]] == 1], Y[stump_cuts[features[idx]] == 1], rule.right, depth + 1)]
            pool.starmap(self.cut, args)
            pool.close()
            pool.join()
            # self.cut(X[stump_cuts[features[idx]]==-1], Y[stump_cuts[features[idx]]==-1], rule.left)
            # print(X[stump_cuts[features[idx]] == 1][features[idx]], Y[stump_cuts[features[idx]] == 1])
            # self.cut(X[stump_cuts[features[idx]] == 1], Y[stump_cuts[features[idx]] == 1], rule.right)
        else:
            # no cut took place
            leaf = Node('leaf', index, p.N, depth=depth + 1, label=p.label, probability=p.probability)
            self.leaf_samples += p.N.sum()/self.data.N
            print(leaf, self.leaf_samples)
            self.tree[index] = leaf
        return

    def print(self, **kwargs):
        traversed = {}
        if 'node' in kwargs:
            node = kwargs['node']
        else:
            node = self.tree[1]

        print(node)
        if node.type == 'rule':
            self.print(node=self.tree[node.left])
            self.print(node=self.tree[node.right])
        return

    def classify(self, X, **kwargs):
        if 'node' in kwargs:
            node = kwargs['node']
        else:
            node = self.tree[1]

        if 'min_leaf_size' in kwargs:
            min_leaf_size = kwargs['min_leaf_size']
        else:
            min_leaf_size = 5

        # print(node)

        lc = self.label_category
        samples = node.count.sum()
        if node.type == 'rule' and samples >= min_leaf_size:
            if X[node.feature] <= node.threshold:
                return self.classify(X, node=self.tree[node.left])
            else:
                return self.classify(X, node=self.tree[node.right])
        else:
            return {lc: node.label}, {lc: node.probability}

    def __init__(self, data, **kwargs):
        self.data = data
        self.label_category = data.label_categories[0]
        self.subsets = {}
        self.tree = {}
        self.leaf_samples = 0
        if 'method' in kwargs:
            self.method = kwargs['method']
        else:
            self.method = 'simple'
        self.cut(data.Xs, data.Ys[self.label_category], 1, 0)
        return


class Performance:
    def __init__(self, classifier, **kwargs):
        self.confusion_matrix = {}

        for lc in classifier.data.label_categories:
            z = zeros((len(classifier.data.class_labels[lc]), len(classifier.data.class_labels[lc])), dtype=int)
            self.confusion_matrix[lc] = pd.DataFrame(z, index=classifier.data.class_labels[lc], columns=classifier.data.class_labels[lc], dtype=int)

        if 'data' in kwargs:
            Xs = kwargs['data']
            Ys = kwargs['Y']
        else:
            Xs = classifier.data.Xs
            Ys = classifier.data.Ys

        for i, X in Xs.iterrows():
            if 'adaboost' in kwargs:
                adaboost = kwargs['adaboost']
            classified_label, probability = classifier.classify(X, adaboost=kwargs['adaboost'])
            for lc in classifier.data.label_categories:
                true_label = Ys[lc][i]
                self.confusion_matrix[lc][classified_label[lc]][true_label] += 1

        self.performance_matrix = {}
        for lc in classifier.data.label_categories:
            performance_matrix = pd.DataFrame(index=classifier.data.class_labels[lc],
                                                       columns=['ppv', 'accuracy', 'sensitivity', 'specificity'])
            detected = self.confusion_matrix[lc].sum()
            true = self.confusion_matrix[lc].sum(axis=1)
            total = array(self.confusion_matrix[lc]).sum()
            TP = correctly_detected = diagonal(self.confusion_matrix[lc])
            TN = total - true - detected + correctly_detected
            FP = detected - correctly_detected
            FN = true - correctly_detected
            performance_matrix['ppv'] = positive_predictive_value = TP/(TP+FP)
            performance_matrix['accuracy'] = (TP+TN)/total
            performance_matrix['sensitivity'] = TP/(TP+FN)
            performance_matrix['specificity'] = TN/(TN+FP)
            self.performance_matrix[lc] = performance_matrix

        return


class AdaBoost:
    def __init__(self, data, **kwargs):
        if 'iter' in kwargs:
            self.iterations = kwargs['iter']
        else:
            self.iterations = 10

        if 'eign' in kwargs:
            self.eign = kwargs['eign']
        else:
            self.eign = 10

        if 'f' in kwargs:
            self.f = kwargs['f']
        else:
            self.f = 0

        if 'samples' in kwargs:
            self.sample_size = kwargs['samples']
        else:
            self.sample_size = int(data.N/2)

        self.data = data

        D = pd.Series(1 / data.N * ones(data.N), name='D')

        f = self.f
        self.h = []
        self.alpha = []
        self.ep = []
        self.en = []
        lc = data.label_categories[0]
        N = freq_count(data.Ys[lc])
        P = N / N.sum()
        for t in range(self.iterations):
            print("f=", f, "ITERATION= ", t)
            indices = choices(range(data.N), D, k=self.sample_size)
            data_t = ClassifiedData(data.Xs.loc[indices].reset_index(drop=True), data.Ys.loc[indices].reset_index(drop=True))
            self.h.append(decisionTree(data_t))
            hx = array([x[0][lc] for x in data.Xs.apply(self.h[t].classify, axis=1)])
            ep = freq_count(hx[data.Ys[lc] == 1])[-1] / N[1]
            en = freq_count(hx[data.Ys[lc] == -1])[1] / N[-1]
            self.alpha.append(0.5 * log((P[1]*(1-ep)+P[-1]*(1-en))/(P[1]*(1+f)*ep+P[-1]*(1-f)*en)))
            for i in D[hx == 1][data.Ys[lc] == 1].index.tolist():
                D[i] = D[i] * exp(-self.alpha[t])
            for i in D[hx == -1][data.Ys[lc] == -1].index.tolist():
                D[i] = D[i] * exp(-self.alpha[t])
            for i in D[hx == -1][data.Ys[lc] == 1].index.tolist():
                D[i] = D[i] * (1+f)* exp(self.alpha[t])
            for i in D[hx == 1][data.Ys[lc] == -1].index.tolist():
                D[i] = D[i] * (1-f)* exp(self.alpha[t])

            D = D/D.sum()
            self.ep.append(ep)
            self.en.append(en)
        return

    def classify(self, X, **kwargs):
        if 'adaboost' in kwargs:
            n = kwargs['adaboost']
        else:
            n = len(self.h)

        H = array(self.alpha)
        for i, h in enumerate(self.h):
            H[i] *= h.classify(X)[0][lc]

        return {lc: int(sign(H[0:n].sum()))}, {lc: (H[0:n].sum()/array(self.alpha)[0:n].sum())}

